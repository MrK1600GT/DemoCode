# encoding: utf-8
import logging.config
import openpyxl
import pyodbc as db
import sys
import os 
from sys import platform

logging.config.fileConfig(fname='logging.properties', disable_existing_loggers=False)
logger = logging.getLogger('uploadREPLCHAR')
logger.info('Commandline arguments: ' + str(sys.argv))    
logger.info('Start uploadREPLCHAR')

if platform == 'win32':
    DSNFile = os.getcwd() + "\DATASOURCENAME.dsn"
    h_connection = r'FILEDSN=' + DSNFile

elif platform == 'os400':
    h_connection = 'DSN=*LOCAL'

def main():
    dbSchema0 = 'LIBRARY'
    dbTable0 = 'REPLCHAR'
    connection = db.connect(h_connection)
    c0 = connection.cursor()
    sqlStmt0 = "delete from {}.{}".format(dbSchema0, dbTable0)
    logger.info(sqlStmt0)
    c0.execute(sqlStmt0)
    connection.commit()
    
    path = "REPLCHAR.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_rows = sheet_obj.max_row

    for i in range(2, max_rows + 1):
        wert = []
    
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row = i, column = j)
            wert.append(cell_obj.value)
    
        if wert[0] != None:
            sqlStmt0 = "insert into {}.{} values('{}','{}','{}','{}')".format(dbSchema0, dbTable0, \
            wert[0].rstrip(), wert[1].rstrip(), wert[2], wert[3])
            logger.info(sqlStmt0)
            c0.execute(sqlStmt0)
            connection.commit()
    c0.close()
    
    logger.info('End uploadREPLCHAR')

def destroy():
    pass

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt as e:
        logger.error(e)
        destroy()
    except Exception as e:
        logger.error(e)
        destroy()